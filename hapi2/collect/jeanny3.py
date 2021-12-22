# -*- coding: utf-8 -*-

import os
import re
import sys
import csv      
import json
import copy
import shutil
import hashlib
import uuid as uuidmod
import itertools as it
from datetime import date, datetime
from warnings import warn,simplefilter

simplefilter('always', UserWarning) # make warnings display always

# Adding a JSON-serialization to the Collection type
from json import JSONEncoder
def _default(self, obj):
    return getattr(obj.__class__, "export_to_json", _default.default)(obj)
_default.default = JSONEncoder().default
JSONEncoder.default = _default

# Simple dictionaty-based DBMS for the data-mining.
# This DBMS inherits a shema-free database ideology (like MongoDB for instance).
# Each document in the collection has it's own unique ID number.
       
# Structure of a collection:
#
# Document = {'__ID__':Integer, '__DATA__':Dictionary} : Dictionary
# Collection = [Document1, Document2, ...]

# Filtering:
# >> Col = Collection(Col1)
# >> IDs = Col.IDs(filter='var['a']+var['b']>10') # get IDs with condition 
# >> Data = Col.get(IDs)
# This way of defining filters must be avoided in the release because of 
# the obvious vulnerabilities!

# VERSION 3.0

#from collections import OrderedDict # doesn't help a lot
#from addict import Dict # takes less memory than OrderedDict

# !!! write more efficient engine on the top of built-in sqlite module?
# https://pypi.python.org/pypi/sqlite-schemaless/0.1.2
# http://yserial.sourceforge.net/
# Good idea, but what to do about open data structure? We need a version control 
# layer on the top of the database engine!

__version__ = '3.0'
print('jeanny, Ver.'+__version__)

FILENAME_ID = '$FILENAME$' # parameter defining the filename
ITEM_ID = '$UUID$' # id that uniquely identifies each item (used in redistribution of items between collections)

# stub for Python 3: redefining unicode
try:
    unicode
except NameError:
    unicode = str # in Python 3 str is unicode

def uuid():
    # http://stackoverflow.com/questions/2759644/python-multiprocessing-doesnt-play-nicely-with-uuid-uuid4    
    return str(uuidmod.UUID(bytes=os.urandom(16), version=4))    
    
# https://stackoverflow.com/questions/11875770/how-to-overcome-datetime-datetime-not-json-serializable-in-python
def json_serial(obj):
    """JSON serializer for objects not serializable by default json code"""
    if isinstance(obj, (datetime, date)):
        serial = obj.isoformat()
        return serial
    raise TypeError ("Type %s not serializable" % type(obj))

def is_identifier(token):    
    return re.match('^[a-zA-Z_][\w_]*$',token)
    
def process_exp(raw_expression): # think about more proper name
    # change this ('a**2+2') to this ('var["a"]**2+2')
    import shlex
    tokens = list(shlex.shlex(raw_expression))    
    for i,token in enumerate(tokens):
        if is_identifier(token): tokens[i] = 'var["%s"]'%token
    return ''.join(tokens)

#class TabObject:
#    """
#    Class for string representation.
#    Used for the output of tabulate function.
#    """
#    def __init__(self,st):
#        self.__string__ = st
#    def tostr(self):
#        return self.__string__
        
class Collection:
    
    #ID_REP = '$id'

    def __init__(self,**argv):
        self.initialize(**argv)
        
    def __getitem__(self,ID):
        return self.getitem(ID)
        
    def __iter__(self):
        return iter(self.getitems())
        
    def __len__(self):
        return len(self.__dicthash__)

    def export_to_json(self):
        """
        Adding a serialization to the Collection type (see above).
        """
        if self.__path__ in {'',None}:
            warn('The collection is not actually saved')
        return {
            '__class__':'Collection',
            '__name__':self.__name__,
            '__type__':self.__type__,
            '__path__':self.__path__}
                
    def clear(self): # synonym for __init__ (TODO: get rid of it)
        self.initialize() 
        # XXX: use custom method instead of __init__
        # in order not to cause conflicts with inherited classes
        
    def __repr__(self):
        #buf = '%s(%s): %s\n'%(self.__name__,self.__type__,self.__path__)
        #buf += json.dumps(self.keys(),indent=2)
        #print(buf)
        #return buf
        #print(self.tabulate())
        #return self.tabulate(raw=True)
        ids = self.ids()
        buf = self.tabulate(raw=True,IDs=ids[:20])
        if len(ids)>20: buf+='\n...'
        return buf

    def initialize(self,path=None,type=None,name='Default',**argv):
        self.maxid = -1
        self.order = [] # order of columns (optional)
        if path:
            if not type: raise Exception('Collection type is not specified')
            if type=='csv':
                self.import_csv(path,**argv)
            elif type=='folder':                
                self.import_folder(path,**argv)
            elif type=='jsonlist':
                self.import_json_list(path,**argv)
            elif type=='xlsx':
                self.import_xlsx(path,**argv)
            elif type=='fixcol':
                self.import_fixcol(path,**argv)
            else:
                raise Exception('Unknown type: %s'%type)
            self.__type__ = type
            self.__path__ = path
            self.__name__ = name
        else:
            self.__dicthash__ = {}
            self.__name__ = ''
            self.__path__ = './'
            self.__type__ = '__init__'
            #self.__dicthash__ = OrderedDict()
            #self.__dicthash__ = Dict() # !!! ATTENTION !!! when fetching non-existing item, creates it => non-standard behaviour
            
    def export(self,path=None,type=None,**argv):
        """
        Export the collection using several available strategies.
        """
        if path is None: path = self.__path__
        if type is None: type = self.__type__
        if type=='csv':
            self.export_csv(path,**argv)
        elif type=='folder':                
            self.export_folder(path,**argv)
        elif type=='jsonlist':
            self.export_json_list(path,**argv)
        elif type=='xlsx':
            self.export_xlsx(path,**argv)
        else:
            raise Exception('Unknown type: %s'%type)        
        
    #def __deepcopy__(self):
    #    """
    #    Function override for the deep copy.
    #    https://stackoverflow.com/questions/1500718/what-is-the-right-way-to-override-the-copy-deepcopy-operations-on-an-object-in-p
    #    """
    #    pass
    
    def copy(self,name='Default',path=''): # CAN BE BUGGY!!
        """
        Copy the collection using the deep copy feature.
        https://www.geeksforgeeks.org/copy-python-deep-copy-shallow-copy/
        """
        col = copy.deepcopy(self)
        col.__name__ = name
        col.__path__ = path # don't use the parent's path by default
        return col

    def setorder(self,order):
        self.order = order
        
    def setfloatfmt(self,floatfmt):
        self.floatfmt = floatfmt
            
    def getitem(self,ID):
        #if type(ID) in [list,tuple]:
        #    if len(ID)>1:
        #        raise Exception('ID must be aither scalar or list of 1 element')
        #    ID = ID[0]
        if ID not in self.__dicthash__:
            #raise Exception('no such ID in __dicthash__: %s'%ID)
            raise KeyError('no such ID in __dicthash__: %s'%ID) # I think that this will mess up the workflow
            #return None
        return self.__dicthash__[ID]
    
    def getitems(self,IDs=-1,mode='strict'):
        """
        Empty IDs must lead to the empty item list!
        Modes: strict,greedy,silent
        """
        if IDs == -1:
            IDs = self.ids()
        buffer = []
        for ID in IDs:
            # some dictionaries create item when it is not found,
            # so do explicit check on item's existence
            if ID not in self.__dicthash__:
                if mode=='strict':
                    raise Exception('no such ID in __dicthash__: %s'%ID)
                elif mode=='silent':
                    continue
                elif mode=='greedy':
                    buffer.append(None)
            else:
                buffer.append(self.__dicthash__[ID])
        return buffer
            
    def getfreeids(self,n):
        # Generate n IDs which don't exist in the collection.
        # TODO: optimize
        idmin = self.maxid + 1
        self.maxid += n # REMOVE THIS FROM HERE
        return list(range(idmin,idmin+n)) # gives error in Python 3 without the list() wrapper
        
    def shuffle(self,n,IDs=-1): # CHANGE NAME!!!
        """
        Shuffle the collection using the round-robin strategy.
        """
        if IDs==-1:
            IDs = self.ids()
        lst_indexes = range(len(IDs))
        shuffled_list = []
        for i in range(n):
            subindex = range(i,len(IDs),n)
            shuffled_sublist = [IDs[i] for i in subindex]
            shuffled_list.append(self.subset(shuffled_sublist))
        return shuffled_list

    def ids(self,filter='True',proc=False):
        if proc: # allows filter be much more simple to input
            filter = process_exp(filter) # experimental
        if type(filter)==str: # simple
            expr = eval('lambda var: ' + filter)
        else:
            expr = filter # advanced
        # this nice trick is taken from stack overflow:
        # http://stackoverflow.com/questions/12467570/python-way-to-speed-up-a-repeatedly-executed-eval-statement?lq=1
        id_list = []
        for ID in self.__dicthash__:
            var = self.__dicthash__[ID]
            #try: # risky
            #    flag = expr(var)
            #except KeyError:
            #    flag = False
            flag = expr(var)
            if flag:
                id_list.append(ID)
        return id_list
        
#    def keys(self):
#        # old version
#        keys = set()
#        for ID in self.__dicthash__:
#            keys = keys.union(self.__dicthash__[ID].keys())
#        keys = list(keys); keys.sort()
#        keys = tuple(keys)  
#        return keys
    
    def keys(self):
        # new version, slow but more informative
        keys = {}
        for ID in self.__dicthash__:
            for key in self.__dicthash__[ID].keys():
                if key not in keys:
                    keys[key] = 1
                else:
                    keys[key] += 1
        return keys

#    def keys(self):
#        # new version #2, slightly faster
#        keys_ = {}
#        # map
#        for ID in self.__dicthash__:
#            k = tuple(self.__dicthash__[ID].keys())
#            if k not in keys_:
#                keys_[k] = 1
#            else:
#                keys_[k] += 1
#        # reduce
#        keys = {}
#        for key_tuple in keys_:
#            for key in key_tuple:
#                if key not in keys:
#                    keys[key] = keys_[key_tuple]
#                else:
#                    keys[key] += keys_[key_tuple]
#        return keys

    def subset(self,IDs=-1):
        if IDs==-1:
            IDs = self.ids()
        new_coll = Collection()
        items = [self.__dicthash__[ID] for ID in IDs]
        new_coll.update(items,IDs)
        new_coll.order = self.order
        return new_coll
    
    def cast(self,type_dict,IDs=-1):
        if IDs==-1:
            IDs = self.ids()
        nchanged = 0
        for ID in IDs:
            if ID not in self.__dicthash__:
                raise Exception('no such ID in __dicthash__: %s'%ID)
            var = self.__dicthash__[ID]
            flag_changed = False
            for col in type_dict:
                if not col in var:
                    continue
                flag_changed = True
                tp = type_dict[col]
                if col in var:
                    var[col] = tp(var[col])
            if flag_changed:
                nchanged += 1
        return {'changed':nchanged}
    
    def batch_(self,expr,IDs=-1):
        if IDs==-1:
            IDs = self.ids()
        ##### OPTIMIZE!!!!!! expr is parsed on each iteration
        for ID in IDs:
            if ID not in self.__dicthash__:
                raise Exception('no such ID in __dicthash__: %s'%ID)
            var = self.__dicthash__[ID]
            exec(expr) # very slow!!!
    
    def assign(self,par,expr,IDs=-1):
        """
        Create new parameter and assign 
        some initial value that may depend on
        other parameters within the item.
        ATTENTION: WILL BE DEPRECATED
        """
        if IDs==-1:
            IDs = self.ids()
        if type(expr)==str: # simple
            expr = eval('lambda var: ' + expr)
        for ID in IDs:
            if ID not in self.__dicthash__:
                raise Exception('no such ID in __dicthash__: %s'%ID)
            var = self.__dicthash__[ID]
            var[par] = expr(var)
        #self.__order__.append(par)
        if par not in self.order:
            self.order.append(par)
            
    def assign_(self,expr,IDs=-1): 
        """
        More flexible and convenient version of assign.
        Takes a lambda, function or callable object 
        as an input, and returns a dictionary
        of the type {'a1':val1,'a2':val2,...}
        where a1,a2,... are new/existing fields, and 
        val1, ... are values to be assigned to those fields.
        The expression takes current item as an input.
        The fields may vary depending on the current item.
        """
        if IDs==-1:
            IDs = self.ids()
        if type(expr)==str: # simple
            expr = eval('lambda var: ' + expr)
        for ID in IDs:
            if ID not in self.__dicthash__:
                raise Exception('no such ID in __dicthash__: %s'%ID)
            var = self.__dicthash__[ID]            
            vals = expr(var)
            for par in vals:
                var[par] = vals[par]
        #self.__order__ += list(dct.keys())

    def assign__(self,dct,IDs=-1):  # expr => dct
        """
        More flexible and convenient version of assign.
        New version takes a dictionary of functions 
        as an input, and returns a dictionary
        of the type {'a1':val1,'a2':val2,...}
        where a1,a2,... are new/existing fields, and 
        val1, ... are values to be assigned to those fields.
        The expression takes current item as an input.
        The fields may vary depending on the current item.
        """
        if IDs==-1:
            IDs = self.ids()
        #if type(expr)==str: # simple
        #    expr = eval('lambda var: ' + expr)
        if type(dct) is not dict:
            raise Exception('dictionary is expected at input')
        for ID in IDs:
            if ID not in self.__dicthash__:
                raise Exception('no such ID in __dicthash__: %s'%ID)
            var = self.__dicthash__[ID]            
            #vals = expr(var)
            #for par in vals:
                #var[par] = vals[par]
            for par in dct:
                var[par] = dct[par](var)
        #self.__order__ += list(dct.keys())
        
    def index(self,expr): # ex-"reform"
        """
        !!! ATTENTION !!!
        Reforms core dictionary by assigning another parameter as key.
        New key MUST be defined for all items in collection.
        Don't use this method you are not sure about consequences.
        Parameter expr can be a string or function. 
        __________________________________________
        !!! TODO: in next index implementation add 
        expressions instead of the field  name.
        __________________________________________
        """
        if type(expr) == str:
            new_id_func = eval('lambda var: var["%s"]'%expr)
            #new_id_func = eval('lambda var: %s'%expr) # this is fucking confusing, don't uncomment it anymore please
        else:
            new_id_func = expr # user-supplied function on item
        # check uniqueness of a new index key
        new_id_vals = []
        for ID in self.__dicthash__:
            item = self.__dicthash__[ID]
            new_id_vals.append(new_id_func(item))
        if len(new_id_vals)!=len(set(new_id_vals)):
            raise Exception('new index is not unique')
        # if there is no duplicates, proceed further
        __dicthash__ = self.__dicthash__ # backup dict hash
        self.__dicthash__ = {}
        keys = list(__dicthash__.keys()) # Python 3 has special object dict_keys
        for ID in keys:
            item = __dicthash__[ID]
            self.__dicthash__[new_id_func(item)] = __dicthash__.pop(ID)
        return self
        
    def get(self,ID,colname):
        """
        Get an element from an item with given ID.
        The point-reference is supported in colnames, i.e.
        "a.b" will search for parameter "b" in the object "a".
        Point-referencing can be multiple.
        """
        if ID not in self.__dicthash__:
            raise Exception('ID=%s is not in dicthash'%str(ID))        
        item = self.__dicthash__[ID]
        chain = colname.split('.')
        cur_obj = item[chain[0]]
        for e in chain[1:]:
            #cur_obj = eval('cur_obj.%s'%e) # atrocious (and slow)
            cur_obj = getattr(cur_obj,e) 
        return cur_obj
        
    def getcols(self,colnames,IDs=-1,strict=True,mode=None,
                functions=None,process=None): # get rid of "strict" argument in ver. 4.0
        """
        Extratct columns from collection.
        If parameter "strict" set to true,
        no exception handling is performed.
        The "functions" parameter is a dictionary containing
        the functions on the item. It also should be present in "colnames".
        Another update: now colname can have a properties,
        such as "col.var"
        __ID__ is a special parameter which corresponds to the local __dicthash__ ID.
        """
        # mode options: 'strict', 'silent', 'greedy'   # add this to docstring in ver. 4.0.
        if not mode: mode = 'strict' if strict else 'silent' 
        if not functions: functions = {}
        #print('%s mode'%mode)
        if IDs==-1:
            IDs = self.ids()
        if type(colnames) is str:
            colnames = [colnames]
        elif type(colnames) is not list:
            raise Exception('Column names should be either list or string')
        cols = []
        for colname in colnames:
            if type(colname) not in [str,unicode]:
                raise Exception('Column name should be a string')
            cols.append([])
        for ID in IDs:
            for i,colname in enumerate(colnames):
                if colname == '__ID__':
                    cols[i].append(ID)
                    continue                    
                try:
                    #cols[i].append(self.__dicthash__[ID][colname]) % old
                    if colname not in functions:
                        cols[i].append(self.get(ID,colname)) # "get" this should be a method of item in the next version of Jeanny
                    else:
                        cols[i].append(functions[colname](self.__dicthash__[ID]))
                except (KeyError, AttributeError) as e: 
                    if mode=='strict':
                        raise e
                    elif mode=='silent':
                        pass
                    elif mode=='greedy':
                        cols[i].append(self.__dicthash__[ID].get(colname))
                    else:
                        raise Exception('unknown mode: %s'%mode)
                    #if strict: raise e # old version with "strict" argument
        if process:
            cols = [process(col) for col in cols]
        return cols
        
    def getcol(self,colname,IDs=-1,strict=True,mode='greedy',functions=None): # get rid of "strict" argument in ver. 4.0
        """
        Wrapper for a single-column call.
        """
        colnames = [colname,]
        return self.getcols(colnames=colnames,IDs=IDs,strict=strict,mode=mode,functions=functions)[0]
        
    def splitcol(self,colname,newcols=None):
        if newcols is None:
            vals = self.getitem(next(iter(self.__dicthash__.keys())))[colname]
            newcols = [colname+'_%d'%i for i,_ in enumerate(vals)]
        for i,cname in enumerate(self.order):
            if cname==colname: self.order.pop(i)
        self.order += newcols
        for item in self.getitems():
            vals = item[colname]
            for newcol,val in zip(newcols,vals):
                item[newcol] = val
            del item[colname]
            
    def deletecols(self,colnames):
        if type(colnames) not in [list,tuple]:
            colnames = [colnames]
        for i,cname in enumerate(self.order):
            if cname in colnames: self.order.pop(i)
        for item in self.getitems():
            for colname in colnames:
                if colname in item:
                    del item[colname]
        
    def tabulate(self,colnames=None,IDs=-1,mode='greedy',fmt='simple',file=None,functions=None,raw=False,floatfmt=None): # switched default to "greedy" instead of "strict"
        """
        Supported table formats are:
        
        - "plain"
        - "simple"
        - "grid"
        - "fancy_grid"
        - "pipe"
        - "orgtbl"
        - "jira"
        - "psql"
        - "rst"
        - "mediawiki"
        - "moinmoin"
        - "html"
        - "latex"
        - "latex_booktabs"
        - "textile"

        More info on usage of Tabulate can be found at 
        https://pypi.org/project/tabulate/
        """
                
        try:
            floatfmt = self.floatfmt
        except AttributeError:
            if floatfmt is None: floatfmt = 'f'
                
        if colnames==None:
            #colnames = list(self.keys().keys()) # this will prevent byg in Python3 since {}.keys() return dict_keys object instead of a list
            allkeys = list(self.keys().keys())
            colnames = self.order + list(set(allkeys)-set(self.order))

        def in_notebook():
            # http://stackoverflow.com/questions/15411967/how-can-i-check-if-code-is-executed-in-the-ipython-notebook            
            """
            Returns ``True`` if the module is running in IPython kernel,
            ``False`` if in IPython shell or other Python shell.
            """
            return 'ipykernel' in sys.modules
        
        from tabulate import tabulate as tab # will make my own tabulator in the future
        data = self.getcols(colnames,IDs=IDs,mode=mode,functions=functions)  
        if file:
            with open(file,'w') as f:
                tabstring = tab(zip(*data),colnames,tablefmt=fmt,floatfmt=floatfmt)
                f.write(tabstring)
        else:
            if in_notebook():
                from IPython.core.display import display, HTML
                tabstring = tab(zip(*data),colnames,tablefmt='html',floatfmt=floatfmt)
                display(HTML(tabstring))
            else:
                tabstring = tab(zip(*data),colnames,tablefmt=fmt,floatfmt=floatfmt)
                if raw:
                    return tabstring
                else:
                    print(tabstring)
                #return TabObject(tabstring)
                
    def head(self):
        self.tabulate(IDs=self.ids()[:20])

    def tail(self):
        ids = self.ids()
        self.tabulate(IDs=ids[len(ids)-20:len(ids)])
    
    def update(self,items,IDs=None):
        if type(items) is dict:
            items = [items]
        elif type(items) not in [list,tuple]:
            raise Exception('Items should be either list or tuple')
        if not IDs:
            IDs = self.getfreeids(len(items))
        #if type(IDs) is int:
        #    IDs = [IDs]
        elif type(IDs) is not list:
            raise Exception('Wrong IDs type: %s (expected list or integer)'%type(IDs))
        for ID,item in zip(IDs,items):
            if ID not in self.__dicthash__:
                self.__dicthash__[ID] = {}
            self.__dicthash__[ID].update(item)
        
    def delete(self,IDs):
        for ID in IDs:
            del self.__dicthash__[ID]
        
    def group(self,expr): # TODO: add process_exp
        buffer = {}
        if type(expr)==str:
            expr = eval('lambda var: (' + expr + ')') # simple
        for ID in self.__dicthash__:
            var = self.__dicthash__[ID] 
            group_value = expr(var)
            if group_value not in buffer:
                buffer[group_value] = []
            buffer[group_value].append(ID)
        return buffer
        
    #def stat(self,group_buffer,operations):  # OLD
    #    # USE ITERATORS!!!!!!!!!!
    #    buffer = {}
    #    for key in group_buffer:
    #        buffer[key] = {}
    #        subindex = {}
    #        for ID in group_buffer[key]:
    #            item = self.__dicthash__[ID]
    #            for col in item:
    #                if col in subindex:
    #                    subindex[col].append(item[col])
    #                else:
    #                    subindex[col] = []
    #        for opkey in operations:
    #            buffer[key][opkey] = operations[opkey](subindex)
    #    return buffer
        
    def stat(self,keyname,grpi,valname,map=None,reduce=None,plain=False):  # Taken from Jeanny v.4 with some changes
        """
        Calculate function on index values.
        User must provide:
            -> group index (grpi)
            -> mapper and reducer functions.
        MAPPER: item->value (can be scalar or vector)
        REDUCER: item_dict_array->value (can be scalar or vector)
        Flat: True - return plain stat index, False - return Collection
        """
        if map is None: map = lambda v: v
        if reduce is None: reduce = lambda ee: ee
        group_buffer = grpi
        stat_index = {}
        for index_id in group_buffer:
            ids = group_buffer[index_id]
            items = self.getitems(ids)
            map_values = [map(item) for item in items]
            reduce_value = reduce(map_values)
            stat_index[index_id] = reduce_value
        if plain: # return only stat index
            return stat_index 
        else: # return index-based collection
            col = Collection()
            col.__dicthash__ = {
                key:{keyname:key,valname:stat_index[key]} \
                for key in stat_index}
            col.order = [keyname,valname]
            return col

    def sort(self,colnames,IDs=-1,strict=True,mode='greedy',functions=None): # switched default to "greedy" instead of "strict"
        """
        Return permutation of the input IDs.
        """
        if IDs==-1:
            IDs = self.ids()
        vals = self.getcols(colnames=colnames,IDs=IDs,strict=strict,mode=mode,functions=functions)
        vals = [list(e)+[id] for e,id in zip(zip(*vals),IDs)]
        IDs_res = [e[-1] for e in sorted(vals)]
        return IDs_res

    def join(self,key,col,colnames=None,prefix=''):
        """
        Join column set of external collection (colnames)
        to self using the key, assuming the following conditions:
        1) Self must contain the column named as key.
        2) Col must have the corresponding values of key in its index.
        Key can be either lambda function on item, or a field name
        """
        if type(key)==str:
            key_ = lambda v: v[key]
        else:
            key_ = key # expecting lambda func, returning col's index values
        if colnames is None:
            colnames_empty = True; order_set = set()
        else:
            colnames_empty = False        
        for item in self.getitems():
            k_ = key_(item)
            if k_ not in col.__dicthash__:
                continue # skip to next iteration if key is not found in external col
            external_item = col.__dicthash__[k_]            
            if colnames_empty:
                colnames = list(external_item.keys())
            for colname in colnames:
                colname_ = prefix+colname
                if colname_ in item:
                    raise Exception('Collection item already has "%s" field'%colname)
                item[colname_] = external_item[colname]
                if colnames_empty: order_set.add(colname_)
        if colnames_empty: 
            self.order += order_set
        else:
            self.order += colnames
        
    #def connect(self,IDs,key,col):
    #    """
    #    !!! ATTENTION !!!
    #    This is an experimental join method which is very limited.
    #    Now it can join only by one single parameter which should coincide 
    #    with __dicthash__ key of an external collection.
    #    This function takes IDs of items as an input and returns 
    #    a pair (ids1,ids2), where ids1 is id column for current collection,
    #    and ids2 is id column for external collection.
    #    """
    #    ids1 = []; ids2 = []
    #    for id1 in IDs:
    #        if id1 not in self.__dicthash__:
    #            raise Exception('no such ID in __dicthash__: %s'%id1)
    #        id2 = self.__dicthash__[id1][key]
    #        if id2 in col.__dicthash__:
    #            ids1.append(id1)
    #            ids2.append(id2)
    #    return ids1,ids2

    # IMPORT AND EXPORT (NEXT JEANNY2 VERSION: plugins? inheritance?)
        
    #https://en.wikipedia.org/wiki/Comma-separated_values
    #http://stackoverflow.com/questions/3191528/csv-in-python-adding-an-extra-carriage-return
    #https://docs.python.org/2/library/csv.html

    # =======================================================
    # =================== UNROLL/UNWIND =====================
    # =======================================================
    def unroll(self,keys,IDs=-1):
        """
        Perform the unrolling of the iterable field(s).
        For example, for {'a':[1,2,3],'b':[10,20]} this will give a new collection:
        {'a':1,'b':10}
        {'a':1,'b':20}
        {'a':2,'b':10}
        {'a':2,'b':20}
        {'a':3,'b':10}
        {'a':3,'b':20}        
        The items not containing the keys from the list will be intact. 
        This method does nearly the same as the "unwind" in MongoDB         
        """
        if type(keys) not in [list,tuple]:
            keys = [keys]
        if IDs==-1:
            IDs = self.ids()
        col = Collection(); col.order = self.order
        def to_list(val):
            if type(val) in [str,unicode]: # string is a special case
                return [val]
            elif type(val) in [list,tuple]: # list and tuple don't need further conversion
                return val
            else: # other cases; TODO: should add iterables separately in the future
                return [val]
        for item in self.getitems(IDs):
            # get the keys from the list which are present in the current item
            active_keys = []
            for key in keys:
                if key in item: active_keys.append(key)
            if not active_keys: # no keys at all
                col.update(item)
            else: # some keys have been found
                for vals in it.product(*[to_list(item[key]) for key in active_keys]):
                    new_item = item.copy()
                    #new_item.update({key:val for key,val in zip(active_keys,vals)}) # this doesn't work in earlier Python versions
                    for key,val in zip(active_keys,vals):
                        new_item[key] = val
                    col.update(new_item)
        return col

    # =======================================================
    # ======================= CSV ===========================
    # =======================================================
    
    def import_csv(self,filename,delimiter=';',quotechar='"',header=None,duck=True):
        """
        Reads csv-formatted files in more or less robust way.
        Includes avoiding many parsing errors due to "illegal"
        usage of delimiters and quotes.
        """
        # TODO: use csv.Sniffer to deduce the format automatically
        items = []
        with open(filename,'r') as f:
            reader = csv.reader(f,delimiter=delimiter,quotechar=quotechar)
            if not header:                
                colnames = next(reader) # take the first line as header (even if it's really absent)
            else:
                colnames = header            
            nitems = 0
            for vals in reader:
                nitems += 1
                item = {}
                for colname,val in zip(colnames,vals):
                    if val is None: continue
                    if duck: # duck typing
                        try:
                            val = int(val)
                        except ValueError as e:
                            try:
                                val = float(val)
                            except ValueError as e:
                                pass
                    if type(val) in [str,unicode]: # f..king encoding problems 
                        try:
                            unicode(val)
                        except UnicodeDecodeError:
                            raise Exception('encoding/decoding problems with %s'%val)
                    #if val: item[colname] = val
                    item[colname] = val
                items.append(item)
        self.clear()
        self.update(items)
        self.order = colnames
        return {'nitems':nitems}
        
    def export_csv(self,filename,delimiter=';',quotechar='"',order=[]):
        """
        Writes csv-formatted files in more or less robust way.
        Includes avoiding many parsing errors due to "illegal"
        usage of delimiters and quotes.
        Order contains key names which will go first.
        Order saves from reordering columns in Excel each time 
        the CSV file is generated.
        """
        if not order: order = self.order
        #header = [key for key in self.keys()]
        keys = self.keys(); 
        header = [key for key in order] + \
                 [key for key in keys if key not in order] # ordered keys must go first
        with open(filename,'w') as f:
            writer = csv.writer(f,delimiter=delimiter,
                  quotechar=quotechar,lineterminator='\n',
                  quoting=csv.QUOTE_MINIMAL)
            writer.writerow(header)
            for ID in self.__dicthash__:
                item = self.__dicthash__[ID]
                vals = []
                for colname in header:
                    if colname in item: # what to do with unicode ?????????????
                        #vals.append(str(item[colname])) # gives error if non-ascii characters are encountered
                        vals.append(unicode(item[colname]))
                        #if type(item[colname]) in (str,unicode):# possible conflicts: Molecules vs Sources
                        #    vals.append(item[colname].encode('utf-8')) # => unicode
                        #else:
                        #    vals.append(unicode(item[colname]))
                    else:
                        vals.append('')
                writer.writerow(vals)
                
    # =======================================================
    # ======================= xlsx ==========================
    # =======================================================
    
    def import_xlsx(self,filename):
        """
        Read in the table stored in the Excel file.
        The upper row must contain the column names.
        """
        from openpyxl import Workbook,load_workbook

        # read workbook
        #wb = load_workbook(filename,use_iterators=True) # Python 3: TypeError: load_workbook() got an unexpected keyword argument 'use_iterators'
        wb = load_workbook(filename)
        sheet = wb.worksheets[0]
        rowlist = list(sheet) # can be inefficient when the file is large; better to use iterators in this case

        # get row and column count
        #http://stackoverflow.com/questions/13377793/is-it-possible-to-get-an-excel-documents-row-count-without-loading-the-entire-d
        #row_count = sheet.max_row
        #column_count = sheet.max_column
    
        # get header, i.e. names of the columns
        header = [cell.value for cell in rowlist[0]]
    
        # fill collection with data
        items = []
        nitems = 0
        for row in rowlist[1:]:
            values = [cell.value for cell in row]
            nitems += 1
            item = {}
            for colname,val in zip(header,values):
                #if val: item[colname] = val # !!!!! BUG: if val==0, it will not be recorded!!!
                if val is not None: item[colname] = val 
            #items.append(item)
            # !!!!!! BUG IN EXCEL/openpyxl: if rows are deleted in Excel, they are still there in the sheet, which will result in many empty items
            if item: items.append(item) 

        self.clear()
        self.update(items)
        return {'nitems':nitems}
        
    # =======================================================
    # ===================== Folder ==========================
    # =======================================================
            
    def import_folder(self,dirname,regex='\.json$'):
        filenames = scanfiles(dirname,regex)
        items = []
        for filename in filenames:
            with open(os.path.join(dirname,filename)) as f:
                try:
                    item = json.load(f)
                except:
                    print('ERROR: %s'%filename)
                    raise
                #if FILENAME_ID in item:
                #    raise Exception('%s has a key %s'%(filenames,FILENAME_ID))
                item[FILENAME_ID] = filename
                items.append(item)
        self.clear()
        self.update(items)
    
    def export_folder(self,dirname,ext='json'):
        """
        Updated version with integrity checking
        to prevent overwriting items in the folder
        in the case when there are similar file names.
        """
        if not os.path.isdir(dirname):
            #os.mkdir(dirname) # this doesn't work if there are sub-folders
            os.makedirs(dirname)
        # prepare file name index
        FILENAMES = {}
        for ID in self.__dicthash__:
            item = self.__dicthash__[ID]
            if FILENAME_ID in item:
                filename = item[FILENAME_ID]
                if ext:
                    filename,_ = os.path.splitext(filename)
                    filename += ext if ext[0]=='.' else '.'+ext
            elif ITEM_ID in item:
                filename = item[ITEM_ID] 
                filename += ext if ext[0]=='.' else '.'+ext
            else:
                #filename = str(uuid.UUID(bytes=os.urandom(16),version=4)) 
                filename = uuid() 
                filename += ext if ext[0]=='.' else '.'+ext
            FILENAMES[ID] = filename
        # check this index for integrity
        #if len(set(FILENAMES.keys()))!=len(FILENAMES.keys()):# BUG!!!!!!
        if len(set(FILENAMES.values()))!=len(FILENAMES.values()):
            raise Exception('%s index is not unique'%FILENAME_ID)
        # if everything is OK save the collection
        for ID in self.__dicthash__:
            item = self.__dicthash__[ID].copy()
            if FILENAME_ID in item:
                del item[FILENAME_ID]
            filename = FILENAMES[ID]
            with open(os.path.join(dirname,filename),'w') as f:
                #json.dump(item,f,indent=2) # default "dumper"
                json.dump(item,f,indent=2,default=json_serial) # custom "dumper"
                     
    def update_folder(self,dirname,regex=''):
        """
        Update the Jeanny folder collection with self.
        !!! BOTH COLLECTIONS MUST HAVE THE SAME IDS !!!
        !!! ONE MUST USE THE FILENAME_ID FOR THESE PARAMETERS !!!
        """
        dest_col = Collection()
        if os.path.isdir(dirname):
            dest_col.import_folder(dirname,regex)
            dest_col.index('var["%s"]'%FILENAME_ID)
        for item in self.getitems():
            FILENAME = item[FILENAME_ID]
            if FILENAME in dest_col.__dicthash__.keys():
                dest_item = dest_col.getitem(FILENAME)
                dest_item.update(item)
            else:
                dest_col.update(item)
        dest_col.export_folder(dirname)

#    def import_json_list(self,filename,id=None): # ancient version of list import
#        """
#        Id is the name of id field in the input file.
#        If id=none, then it is assigned automatically.
#        id is only considered if there is no ID_REP parameter in item.
#        """
#        # TODO: optimize
#        import json
#        import number
#        if id:
#            id_name = id
#        else:
#            id_name = ID_REP
#        with open(filename,'r') as f:
#            items = json.load(f)
#        # scan #1
#        existing_ids = []
#        for item in items:
#            if ID_REP in item:
#                if isinstance(item[id_name], numbers.Number):
#                    existing_ids.append(item[id_name])
#        free_ids = list(set(range(max(existing_ids)))-set(existing_ids))
#        free_ids = list(reversed(free_ids))
#        # scan #2
#        for item in items:
#            if id_name in item:
#                ID = item[id_name]
#            else:
#                ID = free_ids.pop()
#            self.update(item,ID)

    # =======================================================
    # ===================== JSON List =======================
    # =======================================================
        
    def import_json_list(self,filename,id=None):
        with open(filename,'r') as f:
            buffer = json.load(f)
        items = []
        for item in buffer:
            items.append(item)
        self.clear()
        self.update(items)
        
    def export_json_list(self,filename):
        buffer = self.getitems(self.ids())
        with open(filename,'w') as f:
            #json.dump(buffer,f,indent=2) # default "dumper"
            json.dump(buffer,f,indent=2,default=json_serial) # custom "dumper"

    # =======================================================
    # ===================== JSON Dicthash ===================
    # =======================================================
            
    #def import_json_dicthash(self,filename):
    #    import json
    #    with open(filename,'r') as f:
    #        self.__dicthash__ = json.load(f)
    #        self.maxid = max(self.__dicthash__.keys())
        
    def export_json_dicthash(self,filename):
        with open(filename,'w') as f:
            #json.dump(self.__dicthash__,f,indent=2) # default "dumper"
            json.dump(self.__dicthash__,f,indent=2,default=json_serial) # custom "dumper"
            
    # =======================================================
    # ============= Fixcol/Parse from string ================
    # =======================================================
    # THINK THIS THROUGH (ctrl+K to comment, ctrl+shift+K to uncomment the whole block)
            
    def import_fixcol(self,filename,ignore=True,substitute=None):
        """
        Create collection from the specially formatted column-fixed file.
        THe file must be supplied in the following format (type can be omitted):
        
        //HEADER
        0 Column0 Type0
        1 Column1 Type1
        ...
        N ColumnN TypeN
        
        //DATA
        0___1___2____.....N______
        .... data goes here ....
        
        Comments are marked with hashtag (#) and ignored.
        
        If ignore set to False, exception is thrown at any conversion problems.
        """               
        TYPES = {'float':float,'int':int,'str':str}
        
        f = open(filename)
    
        # Search for //HEADER section.    
        for line in f:
            if '//HEADER' in line: break
    
        # Scan //HEADER section.     
        HEAD = {}        
        for line in f:
            line = line.strip()
            if not line: continue
            if line[0]=='#': continue
            if '//DATA' in line: break
            vals = [_ for _ in line.split() if _]
            token = vals[0]
            if token in HEAD:
                raise Exception('ERROR: duplicate key was found: %s'%vals[0])
            vtype = TYPES[vals[2]] if len(vals)>2 else str           
            HEAD[token] = {}
            HEAD[token]['token'] = token
            HEAD[token]['name'] = vals[1]
            HEAD[token]['type'] = vtype # vtype
                        
        # Get tokenized mark-up.
        for line in f:
            widths = line.rstrip(); break # readline doesn't work because of the "Mixing iteration and read methods"
        matches = re.finditer('([^_]+_*)',widths)
        tokens = []; names = []
        for match in matches:
            i_start = match.start()
            i_end = match.end()
            token = re.sub('_','',widths[i_start:i_end])
            if token not in HEAD: continue                
            tokens.append(token)
            names.append(HEAD[token]['name'])
            HEAD[token]['i_start'] = i_start
            HEAD[token]['i_end'] = i_end
        #markup = re.findall('([^_]+_*)',widths) # doesn't give indexes
        
        # Scan //DATA section.     
        items = []
        for line in f:
            if line.strip()=='': continue
            if line.lstrip()[0]=='#': continue
            line = line.rstrip()
            #item = {HEAD[token]['name']:HEAD[token]['type'](line[HEAD[token]['i_start']:HEAD[token]['i_end']]) for token in HEAD} # doesn't work in earlier versions of Python
            item = {}
            for token in HEAD:
                try:
                    buf = line[HEAD[token]['i_start']:HEAD[token]['i_end']]
                    val = HEAD[token]['type'](buf)
                except ValueError as e:
                    if not ignore:
                        raise Exception(e)
                    else:
                        val = substitute
                item[HEAD[token]['name']] = val
            items.append(item) 
        self.clear()
        self.setorder(names)
        self.update(items)    
    
    #def parse(buffer,colfix=False,delim=' ',cast=None,duck=True,head=False,skip=0)
    #    return {'head':HAPI_HEADER,'col':col}   

        # """
        # Extract columns from the 2D buffer.
        # Cast is a list of the types (int,str,float)
        # and can be omitted. Cast also can be a single value.
        # Duck=True means try to deduce the type automatically
        # using the "duck typing" approach.
        # If head=True, the first line contains column header.
        # Empty lines are not accounted for in the numbering.
        # If colfix=True, delimiter and line skipping are ignored.
        # """
        # lines = [line.strip() for line in buffer.split('\n') if line.strip()] #eliminate empty lines
        # table = []
        # vars = {'header':[]}
        # import jeanny3 as j
        # for i,line in enumerate(lines):        
            # if i>0 and i<=skip: 
                # print('slip',i)
                # continue # skip the lines after the possible header
            # vals = [e.strip() for e in line.split(delim) if e] # skip empty entries
            # if i==0:
                # if head==True:
                    # vars['header'] = vals
                    # print('head continue')
                    # continue
                # else:
                    # vars['header'] = ['col%d'%(k+1) for k,val in enumerate(vals)]
            # if cast is not None:
                # if type(cast) not in [list,tuple]:
                    # from itertools import cycle
                    # cast = cycle(cast)
                # for val,tp in zip(vals,cast):
                    # val = tp(val)
            # elif duck==True:
                # for i,val in enumerate(vals):
                    # try:
                        # vals[i] = int(vals[i])
                    # except ValueError as e:
                        # try:
                            # vals[i] = float(vals[i])
                        # except ValueError as e:
                            # pass
            # if vals:                
                # table.append(vals)
        # col = j.Collection()
        # col.update([{h:v for h,v in zip(vars['header'],tabline)} for tabline in table])
        # return col        
        
    # def export_colfix(self):
        # pass
            
    # =======================================================
    # ============= XSCDB/HAPI2.0 STUFF =====================
    # =======================================================
    
    def xscdb_lookup_molecule(self,colname,IDs=-1):
        import xscdb
        if not xscdb.VARSPACE['session']:
            xscdb.start()
        lookup = lambda al: xscdb.query(xscdb.Molecule).\
                            join(xscdb.MoleculeAlias).\
                            filter(xscdb.MoleculeAlias.alias.like(al)).first()
        if IDs==-1: 
            IDs = self.ids()
        for ID in IDs:
            v = self.__dicthash__[ID]
            altypes = ['name','csid','cas','acronym']
            res = {}
            for altype in altypes:                
                if altype in v:
                    dbitem = lookup(v[altype])
                    if dbitem is not None: res[altype] = dbitem
            vals = set(res.values())
            if len(vals)>1:
                #print('WARNING: more than one entrie found for %s: %s'%\
                #  ({altype:v[altype] for altype in altypes if altype in v},res)) # doesn't work in earlier versions of Python
                aa = {}
                for altype in altypes:
                    if altype in v: aa[altype] = v[altype]
                print('WARNING: more than one entrie found for %s: %s'%(aa,res))
                v[colname] = tuple(vals)
            else:
                v[colname] = tuple(vals)[0]
                
    # =======================================================
    # ============= Checksums and integrity =================
    # =======================================================
    
    def md5(self,v):# Jeanny4: this must belong to item, not to collection
        return hashlib.md5('%s'%[v[key] for key in sorted(v.keys())]).hexdigest()

    # =======================================================
    # ============= Other unfinished stuff... ===============
    # =======================================================
        
    def import_binary(self,filename):
        pass
        
    def export_binary(self,filename):
        pass

class Tree: # can a collection do that ??????????
    """
    A collection tree used especially for the cluster parallelized computations
    (temporary implementation).
    """
    
    def __init__(self,folders=None,type='folder',regex='\.json'):   
        self.__cols__ = []
        if folders is not None:
            self.read(folders,type=type,regex=regex)
            
    def __repr__(self):
        res = ',\n'.join([str(col) for col in self.__cols__])
        return res
        
    def read(self,folders,type='folder',regex='\.json'):
        for path in folders:
            print('reading collection from %s'%path)
            col = Collection(path=path,type=type,regex=regex)
            self.__cols__.append(col)
                
    def write(self,folders=None,type=None):
        # check pathes first
        for col in self.__cols__:
            if col.__path__ in {'',None}:
                raise Exception('Path should be non-empty for col %s'%col)
        # get folders
        if folders is None:
            folders = []
            for col in self.__cols__:
                folders.append(col.__path__)
        for col,path in zip(self.__cols__,folders):
            print('exporting %s to %s'%(col,path))
            col.export(path=path,type=type)
        
    def assign(self,par,expr):
        for col in self.__cols__:
            col.assign(par,expr)
        
    def assign_(self,expr):
        for col in self.__cols__:
            col.assign_(expr)
            
    def delete(self,**argv):
        for col in self.__cols__:
            col.delete(**argv)
            
    def subset(self,**argv):
        t = Tree()        
        for col in self.__cols__:
            t.__cols__.append(col.subset(**argv))
        return t
            
    def union(self):
        res = Collection()
        for col in self.__cols__:
            res.update(col.getitems())
        return res

class JobManager:
    """
        ncores => number of CPUs
        nnodes=1 => number of nodes
        mempcore=1500 => amount of RAM per core
        name='calc' => default job name
        command=None => command to launch
        walltime='24' => walltime in hours
    """
    def __init__(self,ncores=1,nnodes=1,mempcore=1500,
        name='calc',command=None,walltime='24'):
        self.ncores = ncores
        self.nnodes = nnodes
        self.mempcore = mempcore
        self.name = name
        self.command = command
        self.walltime = walltime
            
class JobManagerSGE(JobManager):
    template = """# /bin/sh 
# ----------------Parameters---------------------- #
#$ -S /bin/sh
#$ -pe mthread {ncores}
#$ -l s_cpu={walltime}:00:00
#$ -l mres={mempcore}M
#$ -cwd
#$ -j y
#$ -N {name}
#$ -o {name}.log
#$ -m bea
#
# ----------------Modules------------------------- #
#module load tools/python2.6-x
module load opt-python
module load intel

source ~/.bashrc
#
# ----------------Your Commands------------------- #
#
echo + `date` job $JOB_NAME started in $QUEUE with jobID=$JOB_ID on $HOSTNAME
echo + NSLOTS = $NSLOTS
#
{command}
#
echo = `date` job $JOB_NAME done
    """
    def __repr__(self):
        return self.template.format(
            ncores=self.ncores,
            nnodes=self.nnodes,
            mempcore=self.mempcore,
            name=self.name,
            command=self.command,
            walltime=self.walltime)
            
class JobManagerSlurm(JobManager):
    def __repr__(self):
        pass

            
# ==============================
# SUPPLEMENTARY HELPER FUNCTIONS
# ==============================

# get names of all files in a given folder
def get_filenames(dirname):
    filenames = [entry for entry in os.listdir(dirname) 
                 if os.path.isfile(os.path.join(dirname,entry))]
    return filenames
    
def get_dirnames(dirname):
    dirnames = [entry for entry in os.listdir(dirname) 
                 if os.path.isdir(os.path.join(dirname,entry))]
    return dirnames

# filter string according to supplied regular expression (PCRE)
def filterstr(lst,regex):
    return [entry for entry in lst if re.search(regex,entry)]

# scan folder for files which obey the given regular expression (PCRE)
def scanfiles(dirname='./',regex=''):
    return filterstr(get_filenames(dirname),regex)
scandir = scanfiles # BACKWARDS COMPATIBILITY!!
    
# scan folder for sub-folders which obey the given regular expression (PCRE)
def scandirs(dirname='./',regex=''):
    return filterstr(get_dirnames(dirname),regex)

def copyfile(srcdir,srcnames,destdir,destnames=None):
    if not destnames: destnames = srcnames
    for srname,destname in zip(srcnames,destnames):
        # unlike copy(), copy2() retains file attributes
        shutil.copy2(os.path.join(srcdir,srname),
                    os.path.join(destdir,destname))
    
# convert old HAPI-formatted table to Collection    
def collect_hapi(LOCAL_TABLE_CACHE,TableName):
    lines = Collection()
    for i in range(LOCAL_TABLE_CACHE[TableName]['header']['number_of_rows']):
        line = {}
        for par in LOCAL_TABLE_CACHE[TableName]['data'].keys():
            line[par] = LOCAL_TABLE_CACHE[TableName]['data'][par][i]
        lines.update(line)
    return lines
    
# ATTENTION!!! BETTER VERSIONS FOR FUNCTIONS FOR WORKING WITH .PAR HITRAN FORMAT 
# ARE GIVEN IN D:\work\Activities\HAPI\EXCEL_INTERFACE\dotpar\dotpar_converter.py
    
# import HITRAN .par file into a collection
"""
-----------------------------------------------------------------------------------------------------------------------------------------------------------------+
par_line                                                                                                                                                         |
-----------------------------------------------------------------------------------------------------------------------------------------------------------------+
281 0.000001901 1.298E-36 1.637E-25.05940.103  672.98580.580.000000      0 0 0 0        0 0 0 0   12  6 0A-      12  6 0A+     935430 5 4 2 2 1 0   200.0  200.0 |
281 0.000002071 2.913E-35 6.874E-25.07050.105   48.61680.670.000000      0 0 0 0        0 0 0 0    3  3 0A-       3  3 0A+     945430 5 4 2 2 1 0    56.0   56.0 |
281 0.000005174 5.096E-36 2.830E-24.05620.101  787.77310.570.000000      0 0 0 0        0 0 0 0   13  6 0A-      13  6 0A+     935430 5 4 2 2 1 0   216.0  216.0 |
281 0.000012976 1.636E-35 3.866E-23.05300.100  911.20010.560.000000      0 0 0 0        0 0 0 0   14  6 0A-      14  6 0A+     935430 5 4 2 2 1 0   232.0  232.0 |
281 0.000014477 9.247E-34 1.411E-22.07300.109   84.21490.660.000000      0 0 0 0        0 0 0 0    4  3 0A-       4  3 0A+     945430 5 4 2 2 1 0    72.0   72.0 |
281 0.000030368 4.412E-35 4.336E-22.04990.099 1043.22630.550.000000      0 0 0 0        0 0 0 0   15  6 0A-      15  6 0A+     935430 5 4 2 2 1 0   248.0  248.0 |
281 0.000057843 9.690E-33 5.998E-21.07280.109  128.68900.650.000000      0 0 0 0        0 0 0 0    5  3 0A-       5  3 0A+     945430 5 4 2 2 1 0    88.0   88.0 |
281 0.000067033 1.019E-34 4.113E-21.04690.097 1183.80910.540.000000      0 0 0 0        0 0 0 0   16  6 0A-      16  6 0A+     935430 5 4 2 2 1 0   264.0  264.0 |
281 0.000140651 2.048E-34 3.375E-20.04390.096 1332.90370.530.000000      0 0 0 0        0 0 0 0   17  6 0A-      17  6 0A+     935430 5 4 2 2 1 0   280.0  280.0 |
281 0.000173303 5.666E-32 1.152E-19.07170.109  182.02350.640.000000      0 0 0 0        0 0 0 0    6  3 0A-       6  3 0A+     945430 5 4 2 2 1 0   104.0  104.0 |
-----------------------------------------------------------------------------------------------------------------------------------------------------------------+
par_line                                                                                                                                                         |
-----------------------------------------------------------------------------------------------------------------------------------------------------------------+
281 0.000282342 3.628E-34 2.442E-19.04100.094 1490.46320.520.000000      0 0 0 0        0 0 0 0   18  6 0A-      18  6 0A+     835430 5 4 2 2 1 0   296.0  296.0 |
281 0.000545084 5.719E-34 1.580E-18.03820.092 1656.43890.510.000000      0 0 0 0        0 0 0 0   19  6 0A-      19  6 0A+     835430 5 4 2 2 1 0   312.0  312.0 |
281 0.003514511 3.319E-30 3.669E-16.06150.104  483.54750.600.000000      0 0 0 0        0 0 0 0   10  3 0A-      10  3 0A+     845430 5 4 2 2 1 0   168.0  168.0 |
281 0.006136062 5.752E-30 1.627E-15.05840.103  580.84530.590.000000      0 0 0 0        0 0 0 0   11  3 0A-      11  3 0A+     835430 5 4 2 2 1 0   184.0  184.0 |
281 0.010200974 8.732E-30 6.323E-15.05540.102  686.84820.580.000000      0 0 0 0        0 0 0 0   12  3 0A-      12  3 0A+     835430 5 4 2 2 1 0   200.0  200.0 |
281  17.8054720 9.543E-22 2.333E-04.07820.114    8.90430.690.000000      0 0 0 0        0 0 0 0    2  0 0A+       1  0 0A+     845430 5 4 2 2 1 0    40.0   24.0 |
281  17.8061608 7.176E-22 3.500E-04.07960.113    8.37110.690.000000      0 0 0 0        0 0 0 0    2  1 0E        1  1 0E      845430 5 4 2 2 1 0    20.0   12.0 |
281  26.7003452 2.890E-21 8.428E-04.07420.112   26.70980.680.000000      0 0 0 0        0 0 0 0    3  0 0A+       2  0 0A+     845430 5 4 2 2 1 0    56.0   40.0 |
281  26.7013756 2.575E-21 1.498E-03.07580.113   26.17730.680.000000      0 0 0 0        0 0 0 0    3  1 0E        2  1 0E      845430 5 4 2 2 1 0    28.0   20.0 |
281  26.7044710 1.623E-21 9.371E-04.07550.107   24.57810.680.000000      0 0 0 0        0 0 0 0    3  2 0E        2  2 0E      845430 5 4 2 2 1 0    28.0   20.0 |
+------------------------------------------------------------------------------------------------------------------------------------------------------------------+
 11    0.072059 2.043E-30 5.088E-12.09190.391 1922.82910.760.003700          0 1 0          0 1 0  4  2  2        5  1  5      5545533321287120 7     9.0   11.0
 11    0.900086 5.783E-35 1.965E-08.07990.352 5613.36960.57-.002400          0 3 0          0 3 0  8  2  7        7  3  4      4442434432287122 8    51.0   45.0
 11    0.895092 3.600E-28 8.314E-09.08010.412 2129.59910.75-.000300          0 1 0          0 1 0  5  3  2        4  4  1      5546633321287120 8    33.0   27.0
 11    0.865759 5.104E-35 2.755E-08.08180.352 5435.41210.680.005300          0 3 0          0 3 0  6  3  3        7  2  6      4445534432257120 7    13.0   15.0
 11    0.768939 7.600E-37 1.894E-13.07820.434 4030.06980.690.000800          1 0 0          0 0 1  4  3  2        3  3  1      4342434432257122 7    27.0   21.0
 11    0.766500 1.393E-37 1.714E-08.05840.298 6655.62940.51-.009100          0 3 0          0 3 0  9  6  4       10  5  5      3342434432297122 9    19.0   21.0
""";

def dotpar_item_to_list(item):
    elements = ('M','I','nu','S','A','gair','gself','E_','nair','dair',
    'Q','Q_','q','q_','err_nu','err_S','err_gair',
    'err_gself','err_nair','err_dair','ref_nu',
    'ref_S','ref_gair','ref_gself','ref_nair',
    'ref_dair','g','g_')
    lst = [item[e] for e in elements]
    return lst

def load_dotpar(line):
    """
    Get the "raw" .par line on input
    and output the dictionary of parameters.
    """
    item = dict(
        M         = int(   line[  0:  2] ),
        I         = int(   line[  2:  3] ),
        nu        = float( line[  3: 15] ),
        S         = float( line[ 15: 25] ),
        A         = float( line[ 25: 35] ),
        gair      = float( line[ 35: 40] ),
        gself     = float( line[ 40: 45] ),
        E_        = float( line[ 45: 55] ),
        nair      = float( line[ 55: 59] ),
        dair      = float( line[ 59: 67] ),
        Q         = str(   line[ 67: 82] ),
        Q_        = str(   line[ 82: 97] ),
        q         = str(   line[ 97:112] ),
        q_        = str(   line[112:127] ),
        err_nu    = int(   line[127:128] ),
        err_S     = int(   line[128:129] ),
        err_gair  = int(   line[129:130] ),
        err_gself = int(   line[130:131] ),
        err_nair  = int(   line[131:132] ),
        err_dair  = int(   line[132:133] ),
        ref_nu    = int(   line[133:135] ),
        ref_S     = int(   line[135:137] ),
        ref_gair  = int(   line[137:139] ),
        ref_gself = int(   line[139:141] ),
        ref_nair  = int(   line[141:143] ),
        ref_dair  = int(   line[143:145] ),
        g         = float( line[145:153] ),
        g_        = float( line[153:160] ),  
    )   
    return item

#def tostr(par,n=None):
#    if type(par) in [str,unicode]:
#        #return '%%%ds'%n%par
#        return '%s'%par
#    elif type(par) == int:
#        return '%%%dd'%n%par
#    elif type(par) == float:
#        return '%%%fd'%n%par

#def dump_dotpar(item):
#    M      = tostr(item['M'])
#    I      = tostr(item['I'])
#    nu     = tostr(item['nu'])
#    S      = tostr(item['S'])
#    A      = tostr(item['A'])
#    gair   = tostr(item['gair'])
#    gself  = tostr(item['gself'])
#    E_     = tostr(item['E_'])
#    nair   = tostr(item['nair'])
#    dair   = tostr(item['dair'])
#    Q      = tostr(item['Q'])
#    Q_     = tostr(item['Q_'])
#    q      = tostr(item['q'])
#    q_     = tostr(item['q_'])
#    err_nu  = tostr(item['err_nu'])
#    err_S   = tostr(item['err_S'])
#    err_gair  = tostr(item['err_gair'])
#    err_gself = tostr(item['err_gself'])
#    err_nair  = tostr(item['err_nair'])
#    err_dair  = tostr(item['err_dair'])
#    ref_nu    = tostr(item['ref_nu'])
#    ref_S     = tostr(item['ref_S'])
#    ref_gair  = tostr(item['ref_gair'])
#    ref_gself = tostr(item['ref_gself'])
#    ref_nair  = tostr(item['ref_nair'])
#    ref_dair  = tostr(item['ref_dair'])
#    g         = tostr(item['g'])
#    g_        = tostr(item['g_'])
#    line = ''.join(['%s' for i in range(28)])%\
#    (
#    M,I,nu,S,A,gair,gself,E_,nair,dair,
#    Q,Q_,q,q_,err_nu,err_S,err_gair,
#    err_gself,err_nair,err_dair,ref_nu,
#    ref_S,ref_gair,ref_gself,ref_nair,
#    ref_dair,g,g_,
#    )
#    if len(line) != 160:
#        raise Exception('internal strlen error for \"\%s"'%line)  
#    return line
        
def import_dotpar(filename):
    col = Collection()
    with open(filename) as f:
        for line in f:
            item = load_dotpar(line)
            col.update(item)
    return col

def import_fixcol(filename):
    col = Collection()
    col.import_fixcol(filename)
    return col

def import_csv(filename):
    col = Collection()
    col.import_csv(filename)
    return col

## export HITRAN .par file from a collection
#def export_dotpar(col,filename):
#    with open(filename,'w') as f:
#        for id in col.ids():
#            item = col.__dicthash__[id] # not compatible with future versions
#            line = dump_dotpar(item)
#            f.write(line+'\n')
                
def create_from_buffer(colname,buffer,rstrip=True,lstrip=True,cast=lambda val:val,comment=[]):
    """
    Create a collection from the buffer with just one column, containing the lines from buffer.
    Usually this function is useful for the line-by-line text processing.
    """
    items = []
    for line in buffer.split('\n'):
        if not line.strip(): continue
        if line.lstrip()[0] in comment: continue
        val = line
        if rstrip: val = val.rstrip()
        if lstrip: val = val.lstrip()
        val = cast(val)
        items.append({colname:val})
    col = Collection()
    col.update(items)
    return col
    
def create_from_buffer_multicol(buffer,cast={},duck=True,header=True,comment=[]):
    """
    Create a multicolumn collection from the buffer, using variable-length space delimiter.
    Colnames must obey the same rule as as an ordinary data line.
    Usually this function is useful for the line-by-line text processing.
    """
    
    # Skip first new lines.
    lines = buffer.split('\n')
    for istart,line in enumerate(lines):
        if line.strip(): break
            
    # Get column names.
    if header:
        names = lines[istart].split()        
        istart = istart+1
    else:
        names = ['c%d'%i for i in range(len(lines[0].split()))]        
    
    # Start parsing values.
    items = []    
    for line in lines[istart:]:
        if not line.strip(): continue
        if line.lstrip()[0] in comment: continue
        vals = line.split()  
        # Typing.
        item = {}
        for val,name in zip(vals,names):       
            if name in cast:  # cast typing
                val = cast[name](val)
            elif duck: # duck typing
                try:
                    val = int(val)
                except ValueError as e:
                    try:
                        val = float(val)
                    except ValueError as e:
                        pass
            else: # do nothing, treat as raw string
                pass
            item[name] = val
        items.append(item)
    
    # Create collection from the list of items.
    col = Collection()
    col.update(items)
    col.order = names
    
    return col
  
#def join_collections(col1,col2,key,dkey,dist=None,map1=None,map2=None):
def create_join_index(col1,key1,col2,key2,dkey,check=lambda v1,v2:True):
    """ 
    Parameter "key" MUST BE a list or tuple of lambda functions !!
    
    Parameter "check" is additional lambda function comparing 
    items of col1 and col2 and giving True/False at the output.
    
    
    UPDATE MANPAGE!!! IT'S OBSOLETE!!!
    
    Join two collections by key and additional conditions.
    
    col1,col2:
        collections to join together.
    
    dkey:
        Joining by key is performed by comparing the distances 
    between the elements of the "key" tuple with the predefined 
    threshold "dkey" tuple. If the exact matching is required, 
    then the distance should by zero when matched, and very large 
    number otherwise (infinity is preferred).
            
    key:
        Key is a lambda function that must return comparable 
    object (e.g. tuple or scalar). Key expects the collection
    item to have certain fields, from which the key tuple is 
    calculated unless the map1 and map2 are not specified. 
    
    map1,map2:
        Lambda functions for col1 and col2 respectively,
    depending on the collection item.
    Must return the dicts of the form {a0:b0,a1:b1...},
    where (ai,bi) is keys:values for the parameters
    expected by the "key" lambda function.
       
    Binary search is used to speed up the comparison. 
    Speed also depends on the value of dkey, smaller is faster.
    
    """    
    # Process dkey to match key in dimension.
    if type(dkey) not in (list,tuple):
        dkey = [dkey]
            
    # Obtain sorted indexes for both collections.
    ids_col1 = col1.sort(lambda v: key_(v,map=map1))
    ids_col2 = col2.sort(lambda v: key_(v,map=map2))
        
    # Run a loop over two mutually aligned sorted collections.
    for id1 in ids_col1:
        # Search for a 
        pass

def join(col1,col2,join_index,colnames1=lambda c:c,colnames2=lambda c:'.%s'%c):
    """
    Join two collection based on the result of the create_join_index function.
    Colnames 1 and 2 are lambda functions renaming colnames, havin the following format:
        colnames = lambda colname: <any name>
    """
    col = Collection()
    col.order = [colnames1(c) for c in col1.order] + \
                [colnames2(c) for c in col2.order]
    for id1,id2 in join_index:
        item = {}
        if id1:
            item1 = col1.getitem(id1)
            for c in item1:
                c_ = colnames1(c)
                item[c_] = item1[c]
        if id2:
            item2 = col2.getitem(id2)
            for c in item2:
                c_ = colnames2(c)
                if c_ in item:
                    raise Exception('column conflict at join: %s'%c_)
                else:
                    item[c_] = item2[c]
        col.update(item)
    return col
